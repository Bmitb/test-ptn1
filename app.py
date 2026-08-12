"""
Pressure Calibration Report Parser
===================================
A Streamlit application that uses Google Gemini AI to extract structured data
from handwritten pressure calibration reports (PDF/Images) and appends the
parsed data into an existing Excel template with two sheets.

Author  : Senior Python Full-Stack Developer
Version : 1.0.0
"""

from __future__ import annotations

import base64
import io
import json
import os
import re
import tempfile
import traceback
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# PAGE CONFIG  (must be the very first Streamlit call)
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Pressure Calibration Parser",
    page_icon="🔬",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown(
    """
<style>
/* ── Google Font ── */
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

/* ════════════════════════════════════════════════════════
   GLOBAL & BACKGROUND
   ════════════════════════════════════════════════════════ */
html, body, [class*="css"] {
    font-family: 'Inter', sans-serif !important;
    color: #dde3f5 !important;
}
.stApp {
    background: linear-gradient(135deg, #0f0f1a 0%, #141428 40%, #0d1b2a 100%) !important;
}

/* ════════════════════════════════════════════════════════
   SIDEBAR
   ════════════════════════════════════════════════════════ */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0d1117 0%, #161b27 100%) !important;
    border-right: 1px solid rgba(255,255,255,0.07) !important;
}
/* All sidebar text */
[data-testid="stSidebar"],
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div,
[data-testid="stSidebar"] label {
    color: #c8d4ee !important;
}

/* ════════════════════════════════════════════════════════
   LABELS & WIDGET HEADERS
   ════════════════════════════════════════════════════════ */
[data-testid="stWidgetLabel"] p,
[data-testid="stWidgetLabel"] span,
[data-testid="stWidgetLabel"],
.stTextInput label,
.stSelectbox label,
.stFileUploader label,
.stNumberInput label,
.stTextArea label,
.stMultiSelect label,
label {
    color: #c8d4ee !important;
    font-weight: 500 !important;
    font-size: 0.88rem !important;
}

/* ════════════════════════════════════════════════════════
   TEXT INPUT  (the white box problem)
   ════════════════════════════════════════════════════════ */
/* Wrapper */
.stTextInput > div,
.stTextInput > div > div,
[data-testid="stTextInput"] > div,
[data-testid="stTextInput"] > div > div {
    background: rgba(20, 26, 48, 0.9) !important;
    border-color: rgba(99,179,237,0.25) !important;
    border-radius: 10px !important;
}
/* Actual <input> element */
.stTextInput input,
[data-testid="stTextInput"] input,
input[type="text"],
input[type="password"],
input[type="number"],
input[type="email"],
input[type="search"] {
    color: #eef1fc !important;
    background: rgba(20, 26, 48, 0.9) !important;
    border-color: rgba(99,179,237,0.25) !important;
    border-radius: 10px !important;
    caret-color: #63b3ed !important;
}
.stTextInput input::placeholder,
input::placeholder {
    color: rgba(180, 195, 230, 0.4) !important;
}
.stTextInput input:focus,
input:focus {
    border-color: rgba(99,179,237,0.7) !important;
    box-shadow: 0 0 0 2px rgba(99,179,237,0.15) !important;
}

/* ════════════════════════════════════════════════════════
   TEXTAREA
   ════════════════════════════════════════════════════════ */
textarea,
.stTextArea textarea {
    color: #eef1fc !important;
    background: rgba(20, 26, 48, 0.9) !important;
    border-color: rgba(99,179,237,0.25) !important;
    border-radius: 10px !important;
}

/* ════════════════════════════════════════════════════════
   SELECTBOX  (the most problematic one)
   ════════════════════════════════════════════════════════ */
/* Outer wrapper */
.stSelectbox > div > div,
[data-testid="stSelectbox"] > div,
[data-testid="stSelectbox"] > div > div,
[data-baseweb="select"],
[data-baseweb="select"] > div {
    background: rgba(20, 26, 48, 0.9) !important;
    border-color: rgba(99,179,237,0.25) !important;
    border-radius: 10px !important;
    color: #eef1fc !important;
}
/* Selected value text */
[data-baseweb="select"] span,
[data-baseweb="select"] div,
[data-baseweb="select"] input,
.stSelectbox [data-baseweb="select"] *,
[data-testid="stSelectbox"] span,
[data-testid="stSelectbox"] div {
    color: #eef1fc !important;
    background: transparent !important;
}
/* Dropdown menu */
[data-baseweb="menu"],
[data-baseweb="popover"],
ul[data-baseweb="menu"],
[role="listbox"],
[role="option"] {
    background: #1a2035 !important;
    border: 1px solid rgba(99,179,237,0.2) !important;
    border-radius: 10px !important;
}
[role="option"],
[data-baseweb="menu"] li,
[data-baseweb="menu"] * {
    color: #c8d4ee !important;
    background: transparent !important;
}
[role="option"]:hover,
[data-baseweb="menu"] li:hover {
    background: rgba(99,179,237,0.12) !important;
    color: #ffffff !important;
}

/* ════════════════════════════════════════════════════════
   CHECKBOX
   ════════════════════════════════════════════════════════ */
.stCheckbox label p,
.stCheckbox label span,
[data-testid="stCheckbox"] label,
[data-testid="stCheckbox"] span,
[data-testid="stCheckbox"] p {
    color: #c8d4ee !important;
    font-size: 0.87rem !important;
}
/* The checkbox box itself */
[data-baseweb="checkbox"] span {
    border-color: rgba(99,179,237,0.5) !important;
    background: rgba(20,26,48,0.9) !important;
}

/* ════════════════════════════════════════════════════════
   FILE UPLOADER — force dark background
   ════════════════════════════════════════════════════════ */
[data-testid="stFileUploader"],
[data-testid="stFileUploaderDropzone"] {
    background: transparent !important;
}
[data-testid="stFileUploader"] > div,
[data-testid="stFileUploader"] section,
[data-testid="stFileUploaderDropzone"],
[data-testid="stFileUploaderDropzone"] > div {
    background: rgba(15, 20, 40, 0.92) !important;
    border: 2px dashed rgba(99,179,237,0.45) !important;
    border-radius: 14px !important;
    transition: all .3s ease !important;
}
[data-testid="stFileUploaderDropzone"]:hover,
[data-testid="stFileUploader"] > div:hover {
    border-color: rgba(99,179,237,0.85) !important;
    box-shadow: 0 0 20px rgba(99,179,237,0.1) !important;
}
[data-testid="stFileUploader"] *,
[data-testid="stFileUploaderDropzone"] * {
    color: #c8d4ee !important;
    background: transparent !important;
}
/* Browse files button */
[data-testid="stFileUploaderDropzone"] button,
[data-testid="stFileUploader"] button {
    background: rgba(99,179,237,0.15) !important;
    color: #63b3ed !important;
    border: 1px solid rgba(99,179,237,0.4) !important;
    border-radius: 8px !important;
    font-weight: 600 !important;
}
[data-testid="stFileUploaderDropzone"] button:hover {
    background: rgba(99,179,237,0.28) !important;
}

/* ════════════════════════════════════════════════════════
   INFO / WARNING / SUCCESS / ERROR ALERTS
   ════════════════════════════════════════════════════════ */
[data-testid="stAlert"],
[data-testid="stAlert"] p,
[data-testid="stAlert"] span,
[data-testid="stAlert"] div,
.stSuccess, .stWarning, .stError, .stInfo {
    color: #eef1fc !important;
}

/* ════════════════════════════════════════════════════════
   CAPTIONS / SECONDARY TEXT
   ════════════════════════════════════════════════════════ */
.stCaption,
[data-testid="stCaptionContainer"],
[data-testid="stCaptionContainer"] p,
[data-testid="stCaptionContainer"] span,
small {
    color: #8a9bc4 !important;
}

/* ════════════════════════════════════════════════════════
   METRIC WIDGET
   ════════════════════════════════════════════════════════ */
[data-testid="stMetricLabel"] p,
[data-testid="stMetricLabel"] {
    color: #8a9bc4 !important;
}
[data-testid="stMetricValue"],
[data-testid="stMetricValue"] *{
    color: #eef1fc !important;
}

/* ════════════════════════════════════════════════════════
   EXPANDER
   ════════════════════════════════════════════════════════ */
details summary p,
details summary span,
.streamlit-expanderHeader p,
.streamlit-expanderHeader {
    color: #a78bfa !important;
    font-weight: 600 !important;
}
[data-testid="stExpander"] summary p {
    color: #a78bfa !important;
}

/* ════════════════════════════════════════════════════════
   MARKDOWN HEADINGS
   ════════════════════════════════════════════════════════ */
.stMarkdown h1, .stMarkdown h2,
.stMarkdown h3, .stMarkdown h4,
h1, h2, h3, h4 {
    color: #eef1fc !important;
}
.stMarkdown p, .stMarkdown li {
    color: #c8d4ee !important;
}
.stMarkdown a { color: #63b3ed !important; }

/* ════════════════════════════════════════════════════════
   DIVIDER
   ════════════════════════════════════════════════════════ */
hr { border-color: rgba(255,255,255,0.08) !important; }

/* ════════════════════════════════════════════════════════
   CUSTOM COMPONENTS
   ════════════════════════════════════════════════════════ */

/* Glass card */
.glass-card {
    background: rgba(255,255,255,0.04);
    border: 1px solid rgba(255,255,255,0.08);
    border-radius: 16px;
    padding: 24px;
    margin-bottom: 20px;
    backdrop-filter: blur(12px);
    transition: border-color .3s ease;
}
.glass-card:hover { border-color: rgba(99,179,237,0.25); }

/* Section header */
.section-header {
    display: flex;
    align-items: center;
    gap: 10px;
    font-size: 1.05rem;
    font-weight: 600;
    color: #63b3ed;
    margin-bottom: 16px;
    padding-bottom: 10px;
    border-bottom: 1px solid rgba(99,179,237,0.2);
}

/* Status badges */
.badge-ok {
    display: inline-block;
    background: linear-gradient(135deg, #22543d, #276749);
    color: #68d391;
    border: 1px solid rgba(104,211,145,.4);
    padding: 3px 12px;
    border-radius: 20px;
    font-size: .78rem;
    font-weight: 600;
}
.badge-error {
    display: inline-block;
    background: linear-gradient(135deg, #742a2a, #9b2c2c);
    color: #fc8181;
    border: 1px solid rgba(252,129,129,.4);
    padding: 3px 12px;
    border-radius: 20px;
    font-size: .78rem;
    font-weight: 600;
}

/* Hero banner */
.hero-banner {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    border: 1px solid rgba(99,179,237,0.2);
    border-radius: 20px;
    padding: 32px;
    text-align: center;
    margin-bottom: 28px;
    position: relative;
    overflow: hidden;
}
.hero-banner::before {
    content: '';
    position: absolute;
    top: -50%; left: -50%;
    width: 200%; height: 200%;
    background: radial-gradient(ellipse at center, rgba(99,179,237,0.06) 0%, transparent 70%);
    pointer-events: none;
}
.hero-title {
    font-size: 2rem;
    font-weight: 700;
    background: linear-gradient(135deg, #63b3ed, #a78bfa, #f687b3);
    -webkit-background-clip: text;
    -webkit-text-fill-color: transparent;
    background-clip: text;
    margin-bottom: 8px;
}
.hero-sub {
    color: rgba(255,255,255,0.75) !important;
    font-size: .95rem;
    font-weight: 400;
}

/* Metric pill */
.metric-pill {
    background: rgba(255,255,255,0.05);
    border: 1px solid rgba(255,255,255,0.1);
    border-radius: 12px;
    padding: 14px 18px;
    text-align: center;
}
.metric-value { font-size: 1.4rem; font-weight: 700; color: #a78bfa !important; }
.metric-label { font-size: .75rem; color: #8a9bc4 !important; margin-top: 2px; }

/* Buttons */
.stButton > button {
    border-radius: 10px !important;
    font-weight: 600 !important;
    transition: all .25s ease !important;
}
.stButton > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 24px rgba(0,0,0,0.4);
}

/* Data editor */
[data-testid="stDataFrame"],
[data-testid="data-grid-canvas"] {
    border-radius: 10px;
    overflow: hidden;
}

/* Scrollbar */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: rgba(255,255,255,.05); }
::-webkit-scrollbar-thumb { background: rgba(99,179,237,.4); border-radius: 3px; }
</style>
""",
    unsafe_allow_html=True,
)


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
GEMINI_MODELS = [
    "gemini-2.5-flash",          # ⭐ Nhanh, rẻ — khuyến nghị
    "gemini-2.5-pro",            # Chính xác nhất
    "gemini-2.5-flash-lite",     # Nhẹ nhất
    "gemini-3-flash-preview",    # Thế hệ mới
    "gemini-3.5-flash",          # Mới nhất ổn định
    "gemini-flash-latest",       # Alias mới nhất
    "gemini-pro-latest",         # Pro mới nhất
]



def fetch_available_models(api_key: str) -> list[str]:
    """
    Query the Gemini API to list all models that support generateContent.
    Uses the new google-genai SDK.
    Returns a sorted list of model IDs (without the 'models/' prefix).
    """
    try:
        from google import genai
        client = genai.Client(api_key=api_key)
        models = [
            m.name.replace("models/", "")
            for m in client.models.list()
            if m.supported_actions and "generateContent" in m.supported_actions
        ]
        # fallback: if supported_actions not present, include all
        if not models:
            models = [
                m.name.replace("models/", "")
                for m in client.models.list()
            ]
        return sorted(models)
    except Exception as e:
        raise ValueError(f"Không thể lấy danh sách model: {e}")

# ─────────────────────────────────────────────────────────────────────────────
# HELPER UTILITIES
# ─────────────────────────────────────────────────────────────────────────────

def _safe_float(val: Any) -> float | None:
    """Convert any value to float, return None on failure."""
    if val is None:
        return None
    try:
        return float(str(val).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _parse_json_response(text: str) -> dict:
    """
    Extract and parse the JSON payload from Gemini's response text,
    even if it accidentally wraps it in markdown code fences.
    """
    text = text.strip()
    # Strip markdown fences
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.MULTILINE)
    text = re.sub(r"\s*```$", "", text, flags=re.MULTILINE)
    text = text.strip()

    # Try direct parse
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass

    # Attempt to grab the outermost {...} block
    match = re.search(r"\{.*\}", text, re.DOTALL)
    if match:
        return json.loads(match.group(0))

    raise ValueError(f"Could not parse JSON from response:\n{text[:500]}")


def _copy_row_style(ws, src_row: int, dest_row: int):
    """Copy cell styles from src_row to dest_row (best-effort)."""
    for col in range(1, ws.max_column + 1):
        src_cell = ws.cell(row=src_row, column=col)
        dest_cell = ws.cell(row=dest_row, column=col)
        if src_cell.has_style:
            dest_cell.font = copy(src_cell.font)
            dest_cell.fill = copy(src_cell.fill)
            dest_cell.border = copy(src_cell.border)
            dest_cell.alignment = copy(src_cell.alignment)
            dest_cell.number_format = src_cell.number_format


def split_pdf_pages(file_bytes: bytes) -> list[bytes]:
    """
    Split a PDF into a list of PNG image bytes, one per page.
    Requires PyMuPDF (fitz). Falls back to pdf2image if fitz is missing.
    Returns a list of PNG bytes (one element per page).
    """
    pages_png: list[bytes] = []
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(stream=file_bytes, filetype="pdf")
        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
            pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0))
            pages_png.append(pix.tobytes("png"))
        return pages_png
    except ImportError:
        pass
    try:
        from pdf2image import convert_from_bytes
        pil_pages = convert_from_bytes(file_bytes, dpi=200)
        for pil_page in pil_pages:
            buf = io.BytesIO()
            pil_page.save(buf, format="PNG")
            pages_png.append(buf.getvalue())
        return pages_png
    except Exception as e:
        raise ValueError(
            f"Không thể render PDF. Cài PyMuPDF: `pip install pymupdf`. Lỗi: {e}"
        )


# ─────────────────────────────────────────────────────────────────────────────

def extract_data_from_document(
    file_bytes: bytes | None = None,
    mime_type: str = "application/pdf",
    api_key: str = "",
    model_name: str = "gemini-2.5-flash",
    image_bytes: bytes | None = None,
) -> dict:
    """
    Send a calibration document or single page image to Google Gemini and return
    the extracted data as a Python dict.
    Accepts either `file_bytes` or `image_bytes`.
    """
    from google import genai
    from google.genai import types as genai_types

    raw_bytes = image_bytes if image_bytes is not None else file_bytes
    if not raw_bytes:
        raise ValueError("Chưa cung cấp dữ liệu file hoặc hình ảnh.")

    client = genai.Client(api_key=api_key)

    # ── Render PDF / Image → list of Part ──────────────────────────────────
    image_parts: list[genai_types.Part] = []

    if mime_type == "application/pdf":
        try:
            pages_png = split_pdf_pages(raw_bytes)
            for page_bytes in pages_png[:5]:
                image_parts.append(
                    genai_types.Part.from_bytes(data=page_bytes, mime_type="image/png")
                )
        except Exception as e:
            raise ValueError(f"Không thể phân tách trang PDF: {e}")
        if not image_parts:
            raise ValueError("Không trích xuất được trang nào từ PDF.")
    else:
        # Image file or pre-rendered PDF page (png/jpg/jpeg)
        effective_mime = "image/png" if mime_type.startswith("application/pdf") else mime_type
        image_parts.append(
            genai_types.Part.from_bytes(data=raw_bytes, mime_type=effective_mime)
        )

    # ── Call Gemini API with automatic retry & fallback ───────────────────
    import time
    contents = [EXTRACTION_PROMPT] + image_parts
    config = genai_types.GenerateContentConfig(
        temperature=0.0,
        max_output_tokens=4096,
    )

    # Candidate models to attempt if the primary one is 503 overloaded
    fallback_models = [model_name] + [
        m for m in ["gemini-2.5-flash", "gemini-1.5-flash", "gemini-2.5-pro", "gemini-flash-latest"]
        if m != model_name
    ]

    last_err = None
    for attempt_model in fallback_models:
        for attempt in range(3):  # 3 retries per model
            try:
                response = client.models.generate_content(
                    model=attempt_model,
                    contents=contents,
                    config=config,
                )
                result_text = response.text
                return _parse_json_response(result_text)
            except Exception as e:
                err_str = str(e)
                last_err = e
                # Check for 503 (server overloaded) or 429 (rate limit)
                if "503" in err_str or "UNAVAILABLE" in err_str or "429" in err_str or "RESOURCE_EXHAUSTED" in err_str:
                    wait_time = (attempt + 1) * 2
                    time.sleep(wait_time)
                    continue
                else:
                    # Other non-retryable errors -> break retry loop and try next fallback model
                    break

    raise ValueError(
        f"Gemini API error (đã tự động thử lại nhưng máy chủ Google đang quá tải 503): {last_err}\n"
        f"💡 Gợi ý: Chọn model khác ở Sidebar (ví dụ: gemini-1.5-flash hoặc gemini-2.5-pro) hoặc thử lại sau 1-2 phút."
    )





# ─────────────────────────────────────────────────────────────────────────────
# CORE FUNCTION 2: EXCEL APPEND
# ─────────────────────────────────────────────────────────────────────────────

def append_to_excel(excel_path: str | Path, extracted_data: dict) -> bytes:
    """
    Append the extracted calibration data to the two sheets of the Excel
    template and return the modified workbook as bytes.

    Parameters
    ----------
    excel_path     : str or Path
        Path to the existing Excel template file.
    extracted_data : dict
        Validated extraction dict (matching the AI JSON schema).

    Returns
    -------
    bytes
        The in-memory workbook bytes ready for download.

    Raises
    ------
    ValueError
        When required sheets are not found.
    """
    wb = load_workbook(excel_path)

    # ── Resolve sheet references (by index or name) ────────────────────────
    if len(wb.sheetnames) < 1:
        raise ValueError("Excel template has no sheets.")

    ws1 = wb.worksheets[0]   # Sheet 1 — DANH MỤC HIỆU CHUẨN
    ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.create_sheet("Sheet2")

    # ── Unpack data ─────────────────────────────────────────────────────────
    gcn_so          = extracted_data.get("gcn_so", "")
    ma_id           = extracted_data.get("ma_id", "")
    ten_uut         = extracted_data.get("ten_uut", "")
    khach_hang      = extracted_data.get("khach_hang", "")
    nguoi_thuc_hien = extracted_data.get("nguoi_thuc_hien", "")
    ngay_hc         = extracted_data.get("ngay_hc", "")
    ket_qua         = extracted_data.get("ket_qua", "OK")
    tem_hc          = extracted_data.get("tem_hc", "")
    ngay_ke_tiep    = extracted_data.get("ngay_ke_tiep", "")
    tb_chuan_1      = extracted_data.get("tb_chuan_1", "")
    don_vi          = extracted_data.get("don_vi", "")
    range_min       = extracted_data.get("range_min", None)
    range_max       = extracted_data.get("range_max", None)
    points          = extracted_data.get("points", [])

    # ────────────────────────────────────────────────────────────────────────
    # SHEET 1 — append a single row
    # ────────────────────────────────────────────────────────────────────────
    next_row_s1 = ws1.max_row + 1

    # Try to copy style from last data row (if it exists & is not header)
    if ws1.max_row >= 2:
        _copy_row_style(ws1, ws1.max_row, next_row_s1)

    row_s1 = [
        gcn_so,           # Col 1
        ma_id,            # Col 2
        ma_id,            # Col 3 — Mã số nhận dạng (same as Mã ID)
        ten_uut,          # Col 4
        khach_hang,       # Col 5
        "",               # Col 6 — Phiếu YCCV (empty)
        nguoi_thuc_hien,  # Col 7
        "DLVN76",         # Col 8 — P.pháp HC (hardcoded default)
        ngay_hc,          # Col 9
        ket_qua,          # Col 10
        tem_hc,           # Col 11
        ngay_ke_tiep,     # Col 12
        tb_chuan_1,       # Col 13
    ]

    for col_idx, value in enumerate(row_s1, start=1):
        ws1.cell(row=next_row_s1, column=col_idx, value=value)

def _get_style_source_row(ws, is_first_point: bool) -> int | None:
    """Return template master style source row (2 for D1, 3 for D2+ if available)."""
    if ws.max_row < 2:
        return None
    if is_first_point:
        return 2
    return 3 if ws.max_row >= 3 else 2


# ─────────────────────────────────────────────────────────────────────────────
# CORE FUNCTION 2: EXCEL APPEND
# ─────────────────────────────────────────────────────────────────────────────

def append_to_excel(excel_path: str | Path, extracted_data: dict) -> bytes:
    """
    Append the extracted calibration data to the two sheets of the Excel
    template and return the modified workbook as bytes.

    Parameters
    ----------
    excel_path     : str or Path
        Path to the existing Excel template file.
    extracted_data : dict
        Validated extraction dict (matching the AI JSON schema).

    Returns
    -------
    bytes
        The in-memory workbook bytes ready for download.

    Raises
    ------
    ValueError
        When required sheets are not found.
    """
    wb = load_workbook(excel_path)

    # ── Resolve sheet references (by index or name) ────────────────────────
    if len(wb.sheetnames) < 1:
        raise ValueError("Excel template has no sheets.")

    ws1 = wb.worksheets[0]   # Sheet 1 — DANH MỤC HIỆU CHUẨN
    ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.create_sheet("Sheet2")

    # ── Unpack data ─────────────────────────────────────────────────────────
    gcn_so          = extracted_data.get("gcn_so", "")
    ma_id           = extracted_data.get("ma_id", "")
    ten_uut         = extracted_data.get("ten_uut", "")
    khach_hang      = extracted_data.get("khach_hang", "")
    nguoi_thuc_hien = extracted_data.get("nguoi_thuc_hien", "")
    ngay_hc         = extracted_data.get("ngay_hc", "")
    ket_qua         = extracted_data.get("ket_qua", "OK")
    tem_hc          = extracted_data.get("tem_hc", "")
    ngay_ke_tiep    = extracted_data.get("ngay_ke_tiep", "")
    tb_chuan_1      = extracted_data.get("tb_chuan_1", "")
    don_vi          = extracted_data.get("don_vi", "")
    range_min       = extracted_data.get("range_min", None)
    range_max       = extracted_data.get("range_max", None)
    points          = extracted_data.get("points", [])

    # ────────────────────────────────────────────────────────────────────────
    # SHEET 1 — append a single row
    # ────────────────────────────────────────────────────────────────────────
    next_row_s1 = ws1.max_row + 1

    # Copy style from master data row 2 if available
    if ws1.max_row >= 2:
        _copy_row_style(ws1, 2, next_row_s1)

    row_s1 = [
        gcn_so,           # Col 1
        ma_id,            # Col 2
        ma_id,            # Col 3 — Mã số nhận dạng (same as Mã ID)
        ten_uut,          # Col 4
        khach_hang,       # Col 5
        "",               # Col 6 — Phiếu YCCV (empty)
        nguoi_thuc_hien,  # Col 7
        "DLVN76",         # Col 8 — P.pháp HC
        ngay_hc,          # Col 9
        ket_qua,          # Col 10
        tem_hc,           # Col 11
        ngay_ke_tiep,     # Col 12
        tb_chuan_1,       # Col 13
    ]

    for col_idx, value in enumerate(row_s1, start=1):
        ws1.cell(row=next_row_s1, column=col_idx, value=value)

    # ────────────────────────────────────────────────────────────────────────
    # SHEET 2 — append measurement point rows
    # ────────────────────────────────────────────────────────────────────────
    last_row_s2 = ws2.max_row

    # Add an empty separator row if the sheet already has data rows
    has_data_s2 = last_row_s2 >= 2
    if has_data_s2:
        last_row_s2 += 1  # blank separator row (leave empty)

    for i, pt in enumerate(points):
        dest_row = last_row_s2 + 1 + i

        style_src = _get_style_source_row(ws2, is_first_point=(i == 0))
        if style_src:
            _copy_row_style(ws2, style_src, dest_row)

        point_id = pt.get("point_id", f"D{i+1}")
        p_value  = _safe_float(pt.get("p_value"))
        p_tang   = _safe_float(pt.get("p_tang"))
        p_giam   = _safe_float(pt.get("p_giam"))

        # Col 5/6 — Min/Max: only on first row of each device block
        min_val = range_min if i == 0 else None
        max_val = range_max if i == 0 else None

        row_s2 = [
            f"{gcn_so}{point_id}",  # Col 1  — Mã Phụ
            gcn_so,                  # Col 2  — GCN Số
            ma_id,                   # Col 3  — Mã QL / Mã ID
            don_vi,                  # Col 4  — Đ.vị
            min_val,                 # Col 5  — Min
            max_val,                 # Col 6  — Max
            point_id,                # Col 7  — Điểm HC
            don_vi,                  # Col 8  — Đơn vị P
            p_value,                 # Col 9  — P
            don_vi,                  # Col 10 — Đơn vị Chuẩn P
            p_tang,                  # Col 11 — P c.tăng
            p_giam,                  # Col 12 — P c.giảm
        ]

        for col_idx, value in enumerate(row_s2, start=1):
            ws2.cell(row=dest_row, column=col_idx, value=value)

    # ── Save to in-memory buffer ─────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# CORE FUNCTION 3: BATCH EXCEL APPEND
# ─────────────────────────────────────────────────────────────────────────────

def append_all_to_excel(excel_path: str | Path, data_list: list[dict]) -> bytes:
    """
    Append ALL extracted calibration records (batch) to the two sheets of the
    Excel template and return the modified workbook as bytes.

    Each device occupies exactly 1 row in Sheet 1 and N point-rows + 1 blank
    separator row in Sheet 2.

    Parameters
    ----------
    excel_path : str or Path
        Path to the existing Excel template file.
    data_list  : list[dict]
        List of validated extraction dicts (one per calibration document).

    Returns
    -------
    bytes
        The in-memory workbook bytes ready for download.
    """
    wb = load_workbook(excel_path)

    if len(wb.sheetnames) < 1:
        raise ValueError("Excel template has no sheets.")

    ws1 = wb.worksheets[0]
    ws2 = wb.worksheets[1] if len(wb.worksheets) > 1 else wb.create_sheet("Sheet2")

    for extracted_data in data_list:
        gcn_so          = extracted_data.get("gcn_so", "")
        ma_id           = extracted_data.get("ma_id", "")
        ten_uut         = extracted_data.get("ten_uut", "")
        khach_hang      = extracted_data.get("khach_hang", "")
        nguoi_thuc_hien = extracted_data.get("nguoi_thuc_hien", "")
        ngay_hc         = extracted_data.get("ngay_hc", "")
        ket_qua         = extracted_data.get("ket_qua", "OK")
        tem_hc          = extracted_data.get("tem_hc", "")
        ngay_ke_tiep    = extracted_data.get("ngay_ke_tiep", "")
        tb_chuan_1      = extracted_data.get("tb_chuan_1", "")
        don_vi          = extracted_data.get("don_vi", "")
        range_min       = extracted_data.get("range_min", None)
        range_max       = extracted_data.get("range_max", None)
        points          = extracted_data.get("points", [])

        # ── Sheet 1: one row per device ───────────────────────────────────
        next_row_s1 = ws1.max_row + 1
        if ws1.max_row >= 2:
            _copy_row_style(ws1, 2, next_row_s1)

        row_s1 = [
            gcn_so, ma_id, ma_id, ten_uut, khach_hang, "",
            nguoi_thuc_hien, "DLVN76", ngay_hc, ket_qua,
            tem_hc, ngay_ke_tiep, tb_chuan_1,
        ]
        for col_idx, value in enumerate(row_s1, start=1):
            ws1.cell(row=next_row_s1, column=col_idx, value=value)

        # ── Sheet 2: N point rows + 1 blank separator per device ─────────
        last_row_s2 = ws2.max_row
        if last_row_s2 >= 2:
            last_row_s2 += 1  # blank separator row

        for i, pt in enumerate(points):
            dest_row = last_row_s2 + 1 + i
            style_src = _get_style_source_row(ws2, is_first_point=(i == 0))
            if style_src:
                _copy_row_style(ws2, style_src, dest_row)

            point_id = pt.get("point_id", f"D{i+1}")
            p_value  = _safe_float(pt.get("p_value"))
            p_tang   = _safe_float(pt.get("p_tang"))
            p_giam   = _safe_float(pt.get("p_giam"))
            min_val  = range_min if i == 0 else None
            max_val  = range_max if i == 0 else None

            row_s2 = [
                f"{gcn_so}{point_id}", gcn_so, ma_id, don_vi,
                min_val, max_val, point_id, don_vi, p_value,
                don_vi, p_tang, p_giam,
            ]
            for col_idx, value in enumerate(row_s2, start=1):
                ws2.cell(row=dest_row, column=col_idx, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Build preview DataFrames from extracted data
# ─────────────────────────────────────────────────────────────────────────────

def _build_sheet1_df(data: dict) -> pd.DataFrame:
    row = {
        "GCN Số":            data.get("gcn_so", ""),
        "Mã ID":             data.get("ma_id", ""),
        "Mã số nhận dạng":   data.get("ma_id", ""),
        "Tên UUT":           data.get("ten_uut", ""),
        "Khách hàng":        data.get("khach_hang", ""),
        "Phiếu YCCV":        "",
        "Người thực hiện":   data.get("nguoi_thuc_hien", ""),
        "P.pháp HC":         "DLVN76",
        "Ngày hiệu chuẩn":  data.get("ngay_hc", ""),
        "Kết quả HC":        data.get("ket_qua", "OK"),
        "Tem hiệu chuẩn":   data.get("tem_hc", ""),
        "Ngày HC kế tiếp":  data.get("ngay_ke_tiep", ""),
        "TB Chuẩn 1":        data.get("tb_chuan_1", ""),
    }
    return pd.DataFrame([row])


def _build_sheet2_df(data: dict) -> pd.DataFrame:
    gcn_so    = data.get("gcn_so", "")
    ma_id     = data.get("ma_id", "")
    don_vi    = data.get("don_vi", "")
    range_min = data.get("range_min")
    range_max = data.get("range_max")
    points    = data.get("points", [])

    rows = []
    for i, pt in enumerate(points):
        point_id = pt.get("point_id", f"D{i+1}")
        rows.append({
            "Mã Phụ":         f"{gcn_so}{point_id}",
            "GCN Số":         gcn_so,
            "Mã QL / Mã ID":  ma_id,
            "Đ.vị":           don_vi,
            "Min":            range_min if i == 0 else "",
            "Max":            range_max if i == 0 else "",
            "Điểm HC":        point_id,
            "Đơn vị P":       don_vi,
            "P":              pt.get("p_value"),
            "Đơn vị Chuẩn P": don_vi,
            "P c.tăng":       pt.get("p_tang"),
            "P c.giảm":       pt.get("p_giam"),
        })

    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=SHEET2_HEADERS)


def _df_from_session_sheet1() -> pd.DataFrame:
    return _build_sheet1_df(st.session_state.get("extracted_data", {}))


def _df_from_session_sheet2() -> pd.DataFrame:
    return _build_sheet2_df(st.session_state.get("extracted_data", {}))


# ─────────────────────────────────────────────────────────────────────────────
# HELPER: Apply user edits back into the canonical extracted_data dict
# ─────────────────────────────────────────────────────────────────────────────

def _apply_sheet1_edits(df: pd.DataFrame, data: dict) -> dict:
    if df.empty:
        return data
    row = df.iloc[0]
    data["gcn_so"]          = str(row.get("GCN Số", ""))
    data["ma_id"]           = str(row.get("Mã ID", ""))
    data["ten_uut"]         = str(row.get("Tên UUT", ""))
    data["khach_hang"]      = str(row.get("Khách hàng", ""))
    data["nguoi_thuc_hien"] = str(row.get("Người thực hiện", ""))
    data["ngay_hc"]         = str(row.get("Ngày hiệu chuẩn", ""))
    data["ket_qua"]         = str(row.get("Kết quả HC", "OK"))
    data["tem_hc"]          = str(row.get("Tem hiệu chuẩn", ""))
    data["ngay_ke_tiep"]    = str(row.get("Ngày HC kế tiếp", ""))
    data["tb_chuan_1"]      = str(row.get("TB Chuẩn 1", ""))
    return data


def _apply_sheet2_edits(df: pd.DataFrame, data: dict) -> dict:
    if df.empty:
        return data
    if not df.empty and "Đ.vị" in df.columns and not df["Đ.vị"].empty:
        data["don_vi"] = str(df.iloc[0].get("Đ.vị", data.get("don_vi", "")))
    first_row = df[df["Min"].astype(str).str.strip() != ""]
    if not first_row.empty:
        data["range_min"] = _safe_float(first_row.iloc[0].get("Min"))
        data["range_max"] = _safe_float(first_row.iloc[0].get("Max"))

    points = []
    for _, row in df.iterrows():
        points.append({
            "point_id": str(row.get("Điểm HC", "")),
            "p_value":  _safe_float(row.get("P")),
            "p_tang":   _safe_float(row.get("P c.tăng")),
            "p_giam":   _safe_float(row.get("P c.giảm")),
        })
    data["points"] = points
    return data


def _apply_batch_edits(
    edited_s1: pd.DataFrame,
    edited_s2: pd.DataFrame,
    batch_results: list[dict],
) -> list[dict]:
    """
    Merge user edits from the unified data editors back into the batch_results
    list.  Sheet-1 rows are matched positionally; Sheet-2 rows are grouped by
    GCN S\u1ed1 and matched to the corresponding batch entry.
    """
    updated: list[dict] = []
    for i, data in enumerate(batch_results):
        d = dict(data)

        # ── apply Sheet-1 edits (positional) ─────────────────────────────
        if not edited_s1.empty and i < len(edited_s1):
            row = edited_s1.iloc[i]
            d["gcn_so"]          = str(row.get("GCN S\u1ed1", d.get("gcn_so", "")))
            d["ma_id"]           = str(row.get("M\u00e3 ID", d.get("ma_id", "")))
            d["ten_uut"]         = str(row.get("T\u00ean UUT", d.get("ten_uut", "")))
            d["khach_hang"]      = str(row.get("Kh\u00e1ch h\u00e0ng", d.get("khach_hang", "")))
            d["nguoi_thuc_hien"] = str(row.get("Ng\u01b0\u1eddi th\u1ef1c hi\u1ec7n", d.get("nguoi_thuc_hien", "")))
            d["ngay_hc"]         = str(row.get("Ng\u00e0y hi\u1ec7u chu\u1ea9n", d.get("ngay_hc", "")))
            d["ket_qua"]         = str(row.get("K\u1ebft qu\u1ea3 HC", d.get("ket_qua", "OK")))
            d["tem_hc"]          = str(row.get("Tem hi\u1ec7u chu\u1ea9n", d.get("tem_hc", "")))
            d["ngay_ke_tiep"]    = str(row.get("Ng\u00e0y HC k\u1ebf ti\u1ebfp", d.get("ngay_ke_tiep", "")))
            d["tb_chuan_1"]      = str(row.get("TB Chu\u1ea9n 1", d.get("tb_chuan_1", "")))

        # ── apply Sheet-2 edits (group by GCN S\u1ed1) ────────────────────────────
        gcn = d.get("gcn_so", "")
        if gcn and not edited_s2.empty and "GCN S\u1ed1" in edited_s2.columns:
            device_rows = edited_s2[edited_s2["GCN S\u1ed1"].astype(str) == gcn]
            if not device_rows.empty:
                d["don_vi"] = str(device_rows.iloc[0].get("\u0110.v\u1ecb", d.get("don_vi", "")))
                first_min = device_rows[
                    device_rows["Min"].astype(str).str.strip().replace("nan", "") != ""
                ]
                if not first_min.empty:
                    d["range_min"] = _safe_float(first_min.iloc[0].get("Min"))
                    d["range_max"] = _safe_float(first_min.iloc[0].get("Max"))
                points: list[dict] = []
                for _, pt_row in device_rows.iterrows():
                    points.append({
                        "point_id": str(pt_row.get("\u0110i\u1ec3m HC", "")),
                        "p_value":  _safe_float(pt_row.get("P")),
                        "p_tang":   _safe_float(pt_row.get("P c.t\u0103ng")),
                        "p_giam":   _safe_float(pt_row.get("P c.gi\u1ea3m")),
                    })
                d["points"] = points

        updated.append(d)
    return updated


# ─────────────────────────────────────────────────────────────────────────────
# MAIN STREAMLIT APPLICATION
# ─────────────────────────────────────────────────────────────────────────────

def main():
    # ── Session state initialisation ──────────────────────────────────────
    _ss_defaults: dict = {
        "extracted_data": {},
        "extraction_done": False,
        "excel_bytes": None,
        "processing": False,
        "available_models": GEMINI_MODELS,
        "selected_model": GEMINI_MODELS[0],
        # Batch-specific state
        "batch_results": [],     # list[dict] — one per successfully parsed file
        "batch_errors": [],      # list[{filename, error}] — failed files
        "batch_done": False,
        "batch_excel_bytes": None,
    }
    for _k, _v in _ss_defaults.items():
        if _k not in st.session_state:
            st.session_state[_k] = _v

    # ── SIDEBAR ───────────────────────────────────────────────────────────
    with st.sidebar:
        st.markdown(
            "<div style='text-align:center; padding: 10px 0 20px;'>"
            "<span style='font-size:2.4rem;'>🔬</span><br>"
            "<span style='font-size:1.05rem; font-weight:700; color:#63b3ed;'>"
            "CalibParser AI</span><br>"
            "<span style='font-size:.75rem; color:rgba(255,255,255,.4);'>"
            "Batch Edition · Pressure Calibration Suite</span></div>",
            unsafe_allow_html=True,
        )
        st.divider()

        # ── API Key section ──────────────────────────────────────────────
        st.markdown(
            "<div class='section-header'>🔑 Google AI Configuration</div>",
            unsafe_allow_html=True,
        )

        env_key = os.environ.get("GOOGLE_API_KEY", "")
        use_env = st.checkbox(
            "Load key from environment variable",
            value=bool(env_key),
            help="Uses the `GOOGLE_API_KEY` environment variable if set.",
        )

        if use_env and env_key:
            api_key = env_key
            st.success("✅ API key loaded from environment", icon="🔑")
        else:
            api_key = st.text_input(
                "Google API Key",
                type="password",
                placeholder="AIzaSy...",
                value="" if use_env else "",
            )
            if api_key:
                st.success("API key entered", icon="✅")
            else:
                st.info("Enter your Google AI Studio API key above.", icon="ℹ️")

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Model selection ──────────────────────────────────────────────
        st.markdown(
            "<div class='section-header'>🤖 Model Settings</div>",
            unsafe_allow_html=True,
        )

        # Fetch available models button
        fetch_col, _ = st.columns([3, 1])
        with fetch_col:
            if st.button(
                "🔍 Lấy danh sách model",
                use_container_width=True,
                disabled=not api_key,
                help="Tự động lấy danh sách model đang hoạt động từ API key của bạn",
            ):
                with st.spinner("Đang kết nối Gemini API..."):
                    try:
                        fetched = fetch_available_models(api_key)
                        if fetched:
                            st.session_state.available_models = fetched
                            # keep current selection if still valid
                            if st.session_state.selected_model not in fetched:
                                st.session_state.selected_model = fetched[0]
                            st.success(f"✅ Tìm thấy {len(fetched)} model", icon="🤖")
                        else:
                            st.warning("Không tìm thấy model nào hỗ trợ generateContent.")
                    except Exception as e:
                        st.error(str(e))

        # Dynamic selectbox using fetched list
        current_models = st.session_state.available_models
        cur_idx = (
            current_models.index(st.session_state.selected_model)
            if st.session_state.selected_model in current_models
            else 0
        )
        selected = st.selectbox(
            "Chọn model",
            current_models,
            index=cur_idx,
            help="Nhấn 'Lấy danh sách model' để cập nhật danh sách theo API key của bạn.",
        )
        st.session_state.selected_model = selected

        # Manual override
        custom_model = st.text_input(
            "Hoặc nhập tên model thủ công",
            placeholder="vd: gemini-2.0-flash",
            help="Nhập chính xác tên model nếu không có trong danh sách trên.",
        )
        model_name = custom_model.strip() if custom_model.strip() else selected

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Excel template upload ────────────────────────────────────────
        st.markdown(
            "<div class='section-header'>📊 Excel Template</div>",
            unsafe_allow_html=True,
        )
        excel_file = st.file_uploader(
            "Upload Excel Template (.xlsx)",
            type=["xlsx"],
            help="Upload your existing Excel template (2 sheets). Data will be appended while preserving formatting.",
            key="excel_uploader",
        )

        if excel_file:
            st.success(f"📁 `{excel_file.name}` loaded", icon="✅")
            st.caption(f"Size: {excel_file.size / 1024:.1f} KB")

        # ── Reset batch ──────────────────────────────────────────────────
        if st.session_state.batch_done and st.session_state.batch_results:
            st.divider()
            st.markdown(
                "<div class='section-header'>🔄 Batch Controls</div>",
                unsafe_allow_html=True,
            )
            n_ok  = len(st.session_state.batch_results)
            n_err = len(st.session_state.batch_errors)
            st.markdown(
                f"<div style='font-size:.82rem; color:#8a9bc4; margin-bottom:8px;'>"
                f"✅ <b style='color:#68d391;'>{n_ok}</b> file OK"
                + (f" &nbsp;|&nbsp; ❌ <b style='color:#fc8181;'>{n_err}</b> lỗi" if n_err else "")
                + "</div>",
                unsafe_allow_html=True,
            )
            if st.button("🗑️ Xóa kết quả & bắt đầu lại", use_container_width=True):
                st.session_state.batch_results = []
                st.session_state.batch_errors  = []
                st.session_state.batch_done    = False
                st.session_state.batch_excel_bytes = None
                st.rerun()

        st.divider()
        st.markdown(
            "<div style='text-align:center; font-size:.72rem; color:rgba(255,255,255,.3);'>"
            "Powered by Google Gemini AI<br>&amp; openpyxl · Batch Edition</div>",
            unsafe_allow_html=True,
        )

    # ── HERO BANNER ───────────────────────────────────────────────────────
    st.markdown(
        """
        <div class="hero-banner">
            <div class="hero-title">🔬 Pressure Calibration Parser — Batch Edition</div>
            <div class="hero-sub">
                AI đọc hàng loạt phiếu hiệu chuẩn áp suất viết tay → tổng hợp &amp; xuất toàn bộ vào Excel chỉ 1 click
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ── HƯỚNG DẪN SỬ DỤNG ────────────────────────────────────────────────
    with st.expander("📖  Hướng dẫn sử dụng — click để mở", expanded=False):
        st.markdown(
            """
            <div style='padding:4px 0 8px;'>
            <div style='display:grid; grid-template-columns:1fr 1fr; gap:16px;'>

              <div style='background:rgba(99,179,237,0.07); border:1px solid rgba(99,179,237,0.2);
                          border-radius:12px; padding:16px;'>
                <div style='font-size:1rem; font-weight:700; color:#63b3ed; margin-bottom:10px;'>
                  🔧 Chuẩn bị trước
                </div>
                <p style='color:#c8d4ee; font-size:.88rem; line-height:1.8; margin:0;'>
                  <b style='color:#a78bfa;'>1. API Key:</b> Đăng ký miễn phí tại
                  <a href='https://aistudio.google.com' target='_blank'
                     style='color:#63b3ed;'>aistudio.google.com</a>
                  → vào <i>Get API Key</i> → sao chép key và dán vào ô <b>Google API Key</b> ở sidebar bên trái.<br><br>
                  <b style='color:#a78bfa;'>2. File Excel mẫu:</b> Upload file <code>.xlsx</code>
                  có sẵn 2 sheet (Sheet 1: danh mục, Sheet 2: số liệu đo).
                  Ứng dụng sẽ thêm dòng mới vào cuối mỗi sheet, không xóa dữ liệu cũ.
                </p>
              </div>

              <div style='background:rgba(167,139,250,0.07); border:1px solid rgba(167,139,250,0.2);
                          border-radius:12px; padding:16px;'>
                <div style='font-size:1rem; font-weight:700; color:#a78bfa; margin-bottom:10px;'>
                  ⚡ Các bước thực hiện
                </div>
                <p style='color:#c8d4ee; font-size:.88rem; line-height:1.9; margin:0;'>
                  <b style='color:#63b3ed;'>① Upload phiếu:</b>
                  Kéo thả hoặc click chọn file PDF / ảnh chụp phiếu hiệu chuẩn.<br>
                  <b style='color:#63b3ed;'>② Trích xuất AI:</b>
                  Nhấn nút <b>⚡ Extract Data with AI</b> — Gemini sẽ đọc và trả về dữ liệu có cấu trúc.<br>
                  <b style='color:#63b3ed;'>③ Kiểm tra & sửa:</b>
                  Xem lại 2 bảng dữ liệu, click vào ô bất kỳ để sửa tay nếu AI đọc sai.<br>
                  <b style='color:#63b3ed;'>④ Lưu Excel:</b>
                  Nhấn <b>💾 Save &amp; Append to Excel</b> → tải file đã cập nhật về máy.
                </p>
              </div>

            </div>
            <div style='margin-top:14px; background:rgba(246,135,179,0.07);
                        border:1px solid rgba(246,135,179,0.2); border-radius:10px; padding:14px;'>
              <span style='color:#f687b3; font-weight:600;'>💡 Lưu ý:</span>
              <span style='color:#c8d4ee; font-size:.87rem;'>
                Gemini AI hỗ trợ đọc chữ viết tay tiếng Việt.
                Nếu kết quả sai, hãy chọn model <b style='color:#a78bfa;'>gemini-1.5-pro</b> (chính xác hơn)
                hoặc chụp ảnh phiếu với ánh sáng tốt, độ phân giải cao.
              </span>
            </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    # ── STEP INDICATORS (4 bước) ─────────────────────────────────────────
    _sc1, _sc2, _sc3, _sc4 = st.columns(4)
    _steps = [
        (_sc1, "①", "Upload hàng loạt file"),
        (_sc2, "②", "AI trích xuất tất cả"),
        (_sc3, "③", "Review &amp; chỉnh sửa"),
        (_sc4, "④", "Lưu Excel 1-Click"),
    ]
    for _col, _num, _lbl in _steps:
        with _col:
            st.markdown(
                f'<div class="metric-pill">'
                f'<div class="metric-value">{_num}</div>'
                f'<div class="metric-label">{_lbl}</div>'
                "</div>",
                unsafe_allow_html=True,
            )

    st.markdown("<br>", unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 1 — Batch Upload
    # ─────────────────────────────────────────────────────────────────────
    st.markdown(
        "<div class='glass-card'>"
        "<div class='section-header'>📄 Bước 1 — Upload hàng loạt phiếu hiệu chuẩn</div>"
        "<p style='color:#94a3c0; font-size:.87rem; margin:-6px 0 14px;'>"
        "Kéo thả <b style='color:#c8d4ee;'>nhiều file cùng lúc</b> vào ô bên dưới, "
        "hoặc click <b style='color:#c8d4ee;'>Browse files</b> để chọn. "
        "Hỗ trợ: <b style='color:#63b3ed;'>PDF</b> (nhiều trang), "
        "<b style='color:#63b3ed;'>PNG</b>, <b style='color:#63b3ed;'>JPG/JPEG</b>."
        "</p>",
        unsafe_allow_html=True,
    )

    uploaded_files = st.file_uploader(
        "Drop your calibration reports here",
        type=["pdf", "png", "jpg", "jpeg"],
        accept_multiple_files=True,
        help="Hỗ trợ chọn nhiều file cùng lúc. PDF (nhiều trang), PNG, JPG/JPEG.",
        key="doc_uploader",
        label_visibility="collapsed",
    )

    if uploaded_files:
        n_files = len(uploaded_files)
        total_kb = sum(f.size for f in uploaded_files) / 1024
        st.markdown(
            f"<div style='background:rgba(99,179,237,0.07); border:1px solid rgba(99,179,237,0.15); "
            f"border-radius:10px; padding:12px 16px; margin-top:10px;'>"
            f"<span style='color:#63b3ed; font-weight:600;'>📦 {n_files} file đã chọn</span> "
            f"<span style='color:#8a9bc4; font-size:.85rem;'>— Tổng: {total_kb:.1f} KB</span></div>",
            unsafe_allow_html=True,
        )
        file_rows_html = ""
        for i, f in enumerate(uploaded_files, 1):
            icon = "📄" if "pdf" in f.type else "🖼️"
            file_rows_html += (
                f"<tr><td style='padding:4px 10px; color:#8a9bc4;'>#{i}</td>"
                f"<td style='padding:4px 10px;'>{icon} {f.name}</td>"
                f"<td style='padding:4px 10px; color:#a78bfa;'>{f.size/1024:.1f} KB</td>"
                f"<td style='padding:4px 10px; color:#63b3ed;'>{f.type}</td></tr>"
            )
        st.markdown(
            f"<table style='width:100%; font-size:.83rem; color:#c8d4ee; margin-top:8px; border-collapse:collapse;'>"
            f"<thead><tr style='border-bottom:1px solid rgba(255,255,255,.1);'>"
            f"<th style='padding:4px 10px; text-align:left; color:#8a9bc4; font-weight:500;'>#</th>"
            f"<th style='padding:4px 10px; text-align:left; color:#8a9bc4; font-weight:500;'>Tên file</th>"
            f"<th style='padding:4px 10px; text-align:left; color:#8a9bc4; font-weight:500;'>Kích thước</th>"
            f"<th style='padding:4px 10px; text-align:left; color:#8a9bc4; font-weight:500;'>Loại</th>"
            f"</tr></thead><tbody>{file_rows_html}</tbody></table>",
            unsafe_allow_html=True,
        )

    st.markdown("</div>", unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 2 — Batch AI Extraction with Progress
    # ─────────────────────────────────────────────────────────────────────
    st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
    st.markdown(
        "<div class='section-header'>🤖 Bước 2 — Trích xuất AI hàng loạt</div>",
        unsafe_allow_html=True,
    )

    n_uploaded = len(uploaded_files) if uploaded_files else 0
    if n_uploaded > 1:
        _btn_label = f"🚀 Trích xuất tất cả {n_uploaded} file"
    elif n_uploaded == 1:
        _btn_label = "⚡ Trích xuất file"
    else:
        _btn_label = "🚀 Trích xuất tất cả các file"

    _ext_col, _spacer = st.columns([1, 3])
    with _ext_col:
        extract_btn = st.button(
            _btn_label,
            use_container_width=True,
            type="primary",
            disabled=(not uploaded_files or not api_key),
            key="btn_extract_batch",
        )

    if not api_key:
        st.warning("⚠️ Vui lòng nhập Google API Key ở thanh Sidebar bên trái.", icon="⚠️")
    if not uploaded_files:
        st.info("ℹ️ Upload ít nhất 1 file phiếu hiệu chuẩn ở Bước 1 để bắt đầu.", icon="ℹ️")

    if extract_btn and uploaded_files and api_key:
        _n = len(uploaded_files)
        _batch_results: list[dict] = []
        _batch_errors:  list[dict] = []

        # Count total "units" to process (each PDF page = 1 unit, each image = 1 unit)
        # We do a quick pre-scan to know how many pages each PDF has
        _unit_plan: list[dict] = []  # {"file": uf, "file_bytes": bytes, "mime": str, "page": int|None, "label": str}
        _prescan_slot = st.empty()
        _prescan_slot.info("⏳ Đang quét số trang PDF...", icon="📚")
        for _uf in uploaded_files:
            _fb = _uf.read()
            _mime = _uf.type or "application/octet-stream"
            if "pdf" in _mime or _uf.name.lower().endswith(".pdf"):
                try:
                    _pages = split_pdf_pages(_fb)
                    for _pg_idx, _pg_bytes in enumerate(_pages):
                        _unit_plan.append({
                            "filename": _uf.name,
                            "file_bytes": _pg_bytes,
                            "mime": "image/png",
                            "page_idx": _pg_idx,
                            "total_pages": len(_pages),
                        })
                except Exception as _e:
                    _batch_errors.append({"filename": _uf.name, "error": str(_e)})
            else:
                _unit_plan.append({
                    "filename": _uf.name,
                    "file_bytes": _fb,
                    "mime": _mime,
                    "page_idx": None,
                    "total_pages": 1,
                })
        _prescan_slot.empty()

        _n_units = len(_unit_plan)
        _progress_bar = st.progress(0.0, text="Chuẩn bị...")
        _status_slot  = st.empty()
        _log_slot     = st.empty()
        _log_lines: list[str] = []

        for _ui, _unit in enumerate(_unit_plan):
            _fname   = _unit["filename"]
            _pg_idx  = _unit["page_idx"]
            _tot_pg  = _unit["total_pages"]
            if _pg_idx is not None:
                _lbl = f"➤ File <b>{_fname}</b> — Trang {_pg_idx + 1}/{_tot_pg}"
            else:
                _lbl = f"➤ File <b>{_fname}</b>"

            _progress_bar.progress(_ui / _n_units, text=f"Đang xử lý {_ui+1}/{_n_units}...")
            _status_slot.markdown(
                f"<div style='background:rgba(99,179,237,0.08); border-left:3px solid #63b3ed; "
                f"border-radius:8px; padding:10px 16px; margin:6px 0; font-size:.9rem; color:#c8d4ee;'>"
                f"⚙️ {_lbl}</div>",
                unsafe_allow_html=True,
            )
            try:
                _result = extract_data_from_document(
                    image_bytes=_unit["file_bytes"],
                    mime_type=_unit["mime"],
                    api_key=api_key,
                    model_name=model_name,
                )
                _result["_source_file"] = (
                    f"{_fname} (trang {_pg_idx+1})"
                    if _pg_idx is not None else _fname
                )
                _batch_results.append(_result)
                _n_pts = len(_result.get("points", []))
                _log_lines.append(
                    f"✅ {_lbl} — GCN: <b>{_result.get('gcn_so','?')}</b> "
                    f"| Mã: {_result.get('ma_id','?')} | {_n_pts} điểm đo"
                )
            except Exception as _exc:
                _batch_errors.append({"filename": _lbl, "error": str(_exc)})
                _log_lines.append(
                    f"❌ {_lbl} — Lỗi: {str(_exc)[:120]}"
                )

            _log_html = "".join(
                f"<div style='font-size:.82rem; padding:2px 0; color:#c8d4ee;'>{_ln}</div>"
                for _ln in _log_lines
            )
            _log_slot.markdown(
                f"<div style='background:rgba(0,0,0,.25); border-radius:10px; "
                f"padding:12px 16px; margin-top:8px;'>{_log_html}</div>",
                unsafe_allow_html=True,
            )

        _progress_bar.progress(1.0, text="Hoàn tất!")
        _n_ok  = len(_batch_results)
        _n_err = len(_batch_errors)
        _status_slot.markdown(
            f"<div style='background:rgba(104,211,145,0.1); border-left:3px solid #68d391; "
            f"border-radius:8px; padding:10px 16px; font-size:.9rem; color:#c8d4ee;'>"
            f"🎉 <b>Hoàn tất!</b> Thành công: <b style='color:#68d391;'>{_n_ok}/{_n_units}</b> phiếu"
            + (f" &nbsp;|&nbsp; Lỗi: <b style='color:#fc8181;'>{_n_err}</b>" if _n_err else "")
            + "</div>",
            unsafe_allow_html=True,
        )

        st.session_state.batch_results    = _batch_results
        st.session_state.batch_errors     = _batch_errors
        st.session_state.batch_done       = True
        st.session_state.batch_excel_bytes = None

        if _n_ok > 0:
            _total_pts = sum(len(r.get("points", [])) for r in _batch_results)
            st.success(
                f"✅ Đã trích xuất **{_n_ok}** phiếu (từ **{_n}** file) — tổng cộng **{_total_pts}** điểm đo.",
                icon="🎉",
            )

    st.markdown("</div>", unsafe_allow_html=True)

    # ─────────────────────────────────────────────────────────────────────
    # SECTION 3 — Unified Review & Edit Table
    # ─────────────────────────────────────────────────────────────────────
    if st.session_state.batch_done and st.session_state.batch_results:
        _br = st.session_state.batch_results
        _be = st.session_state.batch_errors

        st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
        st.markdown(
            "<div class='section-header'>✏️ Bước 3 — Xem trước &amp; Chỉnh sửa dữ liệu tổng hợp</div>",
            unsafe_allow_html=True,
        )

        # ── Summary metrics ───────────────────────────────────────────────
        _mc1, _mc2, _mc3, _mc4 = st.columns(4)
        with _mc1:
            st.metric("📦 File thành công", len(_br))
        with _mc2:
            st.metric("🔖 Tổng thiết bị", len(_br))
        with _mc3:
            st.metric("📐 Tổng điểm đo", sum(len(r.get("points", [])) for r in _br))
        with _mc4:
            st.metric("⚠️ File lỗi", len(_be))

        if _be:
            with st.expander(f"⚠️ {len(_be)} file gặp lỗi — click để xem chi tiết"):
                for _err in _be:
                    st.error(f"❌ **{_err['filename']}**: {_err['error']}", icon="🚨")

        st.markdown("---")

        with st.expander("🔍 Xem dữ liệu JSON thô của tất cả file"):
            st.json(_br)

        st.markdown("---")

        # ── SHEET 1 UNIFIED EDITOR ────────────────────────────────────────
        st.markdown(
            "<div style='font-size:1rem; font-weight:600; color:#c8d4ee; margin-bottom:6px;'>"
            "📋 Sheet 1 — Thông tin chung (mỗi hàng = 1 phiếu hiệu chuẩn)</div>",
            unsafe_allow_html=True,
        )
        st.caption(
            "Mỗi hàng tương ứng với 1 phiếu. Index (#) = thứ tự file. "
            "Click vào ô bất kỳ để chỉnh sửa trực tiếp."
        )
        df_all_s1 = pd.concat(
            [_build_sheet1_df(d) for d in _br], ignore_index=True
        )
        edited_s1 = st.data_editor(
            df_all_s1,
            use_container_width=True,
            num_rows="fixed",
            hide_index=False,
            key="batch_editor_sheet1",
            column_config={
                "GCN Số":           st.column_config.TextColumn("GCN Số", width="medium"),
                "Mã ID":            st.column_config.TextColumn("Mã ID", width="medium"),
                "Mã số nhận dạng":  st.column_config.TextColumn("Mã số nhận dạng", width="medium"),
                "Tên UUT":          st.column_config.TextColumn("Tên UUT", width="small"),
                "Khách hàng":       st.column_config.TextColumn("Khách hàng", width="medium"),
                "Phiếu YCCV":       st.column_config.TextColumn("Phiếu YCCV", width="small"),
                "Người thực hiện":  st.column_config.TextColumn("Người thực hiện", width="medium"),
                "P.pháp HC":        st.column_config.TextColumn("P.pháp HC", width="small"),
                "Ngày hiệu chuẩn": st.column_config.TextColumn("Ngày HC", width="medium"),
                "Kết quả HC":       st.column_config.SelectboxColumn(
                    "Kết quả HC",
                    options=["OK", "FAIL", "N/A"],
                    width="small",
                ),
                "Tem hiệu chuẩn":  st.column_config.TextColumn("Tem HC", width="medium"),
                "Ngày HC kế tiếp": st.column_config.TextColumn("Ngày kế tiếp", width="medium"),
                "TB Chuẩn 1":      st.column_config.TextColumn("TB Chuẩn 1", width="medium"),
            },
        )

        st.markdown("<br>", unsafe_allow_html=True)

        # ── SHEET 2 UNIFIED EDITOR ────────────────────────────────────────
        st.markdown(
            "<div style='font-size:1rem; font-weight:600; color:#c8d4ee; margin-bottom:6px;'>"
            "📊 Sheet 2 — Số liệu đo tổng hợp (tất cả điểm hiệu chuẩn)</div>",
            unsafe_allow_html=True,
        )
        st.caption(
            "Bảng tổng hợp tất cả điểm đo của mọi thiết bị. "
            "Các thiết bị được phân biệt theo cột GCN Số. "
            "Click vào ô bất kỳ để chỉnh sửa."
        )
        df_all_s2 = pd.concat(
            [_build_sheet2_df(d) for d in _br], ignore_index=True
        )
        edited_s2 = st.data_editor(
            df_all_s2,
            use_container_width=True,
            num_rows="dynamic",
            hide_index=False,
            key="batch_editor_sheet2",
            column_config={
                "Mã Phụ":          st.column_config.TextColumn("Mã Phụ", width="large"),
                "GCN Số":          st.column_config.TextColumn("GCN Số", width="medium"),
                "Mã QL / Mã ID":   st.column_config.TextColumn("Mã QL", width="medium"),
                "Đ.vị":            st.column_config.TextColumn("Đ.vị", width="small"),
                "Min":             st.column_config.NumberColumn("Min", format="%.2f"),
                "Max":             st.column_config.NumberColumn("Max", format="%.2f"),
                "Điểm HC":         st.column_config.TextColumn("Điểm HC", width="small"),
                "Đơn vị P":        st.column_config.TextColumn("Đơn vị P", width="small"),
                "P":               st.column_config.NumberColumn("P", format="%.4f"),
                "Đơn vị Chuẩn P":  st.column_config.TextColumn("Đơn vị Chuẩn P", width="small"),
                "P c.tăng":        st.column_config.NumberColumn("P c.tăng", format="%.4f"),
                "P c.giảm":        st.column_config.NumberColumn("P c.giảm", format="%.4f"),
            },
        )

        st.markdown("</div>", unsafe_allow_html=True)

        # ─────────────────────────────────────────────────────────────────
        # SECTION 4 — One-Click Save & Append to Excel
        # ─────────────────────────────────────────────────────────────────
        st.markdown("<div class='glass-card'>", unsafe_allow_html=True)
        st.markdown(
            "<div class='section-header'>💾 Bước 4 — Lưu tất cả vào Excel (1-Click)</div>",
            unsafe_allow_html=True,
        )

        if excel_file is None:
            st.warning(
                "⚠️ Chưa upload file Excel mẫu. Hãy upload file `.xlsx` (2 sheet) ở Sidebar bên trái.",
                icon="📊",
            )
        else:
            _n_devices = len(_br)
            _save_col, _dl_col = st.columns([1, 1])
            with _save_col:
                save_btn = st.button(
                    f"💾 Lưu tất cả {_n_devices} thiết bị vào Excel",
                    use_container_width=True,
                    type="primary",
                    key="btn_save_all",
                )

            if save_btn:
                # Merge edits from unified tables back into batch_results
                _updated_batch = _apply_batch_edits(edited_s1, edited_s2, _br)
                st.session_state.batch_results = _updated_batch

                with st.spinner(
                    f"⚙️ Đang ghi {_n_devices} thiết bị vào Excel — vui lòng chờ…"
                ):
                    try:
                        with tempfile.NamedTemporaryFile(
                            suffix=".xlsx", delete=False
                        ) as _tmp:
                            _tmp.write(excel_file.read())
                            _tmp_path = _tmp.name

                        _excel_bytes = append_all_to_excel(_tmp_path, _updated_batch)
                        st.session_state.batch_excel_bytes = _excel_bytes

                        try:
                            os.unlink(_tmp_path)
                        except OSError:
                            pass

                        _total_pts_saved = sum(
                            len(r.get("points", [])) for r in _updated_batch
                        )
                        st.success(
                            f"✅ Đã ghi **{_n_devices}** thiết bị "
                            f"(**{_total_pts_saved}** điểm đo) vào Excel thành công!",
                            icon="🎉",
                        )
                    except Exception as _exc:
                        st.error(f"❌ Lỗi khi ghi Excel: {_exc}", icon="🚨")
                        with st.expander("📋 Full traceback"):
                            st.code(traceback.format_exc(), language="python")

            # ── Download button ──────────────────────────────────────────
            if st.session_state.batch_excel_bytes:
                _ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
                _filename = f"CalibReport_Batch_{_n_devices}thietbi_{_ts}.xlsx"

                with _dl_col:
                    st.download_button(
                        label=f"⬇️ Tải file Excel ({_n_devices} thiết bị)",
                        data=st.session_state.batch_excel_bytes,
                        file_name=_filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="secondary",
                    )

                st.info(
                    f"📥 File **`{_filename}`** đã sẵn sàng tải về. "
                    "Tất cả định dạng, công thức và dữ liệu cũ trong file template được giữ nguyên.",
                    icon="✅",
                )

        st.markdown("</div>", unsafe_allow_html=True)

    # ── Footer ────────────────────────────────────────────────────────────
    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown(
        "<div style='text-align:center; color:rgba(255,255,255,.25); font-size:.75rem;'>"
        "CalibParser AI · Batch Edition · Built with Streamlit + Google Gemini + openpyxl"
        "</div>",
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    main()
